from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

def truncate_table_name(table_name, existing_names):
    truncated_name = table_name[:31]
    if truncated_name in existing_names:
        index = 1
        while f"{truncated_name[:27]}_{index}" in existing_names:
            index += 1
        truncated_name = f"{truncated_name[:27]}_{index}"
    return truncated_name

def create_index_sheet(wb, table_names, scripts, tableau_fields, header_style, INDEX_TABLE_STYLE, HEADER_INSTRUCTIONS, HEADER_NAVIGATION, HEADER_DATASETS, HEADER_SCRIPTS, TABLEAU_HEADING):
    index_sheet = wb.create_sheet(title="Index", index=0)
    index_sheet.column_dimensions['A'].width = 40
    index_sheet.column_dimensions['B'].width = 40
    index_sheet.column_dimensions['C'].width = 40
    index_sheet.column_dimensions['D'].width = 15
    index_sheet.column_dimensions['E'].width = 15
    index_sheet.sheet_view.showGridLines = False

    index_sheet.append(["Instructions"])
    index_sheet["A1"].style = header_style
    index_sheet.append([HEADER_INSTRUCTIONS])
    index_sheet.append([HEADER_NAVIGATION])
    index_sheet.append([])

    index_sheet.append([HEADER_DATASETS])
    index_sheet["A5"].style = header_style
    index_sheet.append(["Schema", "Table Name", "Sheet Name"])

    for table_name, table_info in table_names.items():
        schema = table_info['schema']
        sheet_name = table_info['sheet_name']
        row_idx = index_sheet.max_row + 1
        index_sheet.cell(row=row_idx, column=1, value=schema)
        index_sheet.cell(row=row_idx, column=2, value=table_name)
        cell = index_sheet.cell(row=row_idx, column=3, value=sheet_name)
        cell.hyperlink = f"#{sheet_name}!A1"
        cell.style = "Hyperlink"

    max_row = index_sheet.max_row
    table = Table(displayName="DatasetsTable", ref=f"A6:C{max_row}")
    style = TableStyleInfo(name=INDEX_TABLE_STYLE, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    index_sheet.add_table(table)

    index_sheet.append([])

    index_sheet.append([HEADER_SCRIPTS])
    scripts_header_row = index_sheet.max_row
    index_sheet[f"A{scripts_header_row}"].style = header_style
    index_sheet.append(["Workflow Name", "Dataset Mnemonic", "Dataset Version", "Date Modified", "Sheet Name"])

    for i, script in scripts.iterrows():
        sheet_name = f"Script_{i + 1}"
        row_idx = index_sheet.max_row + 1
        index_sheet.append([script["WORKFLOW_NAME"], script["DATA_SET_MNEMONIC"], script["DATA_SET_VERSION"], script["DATE_MODIFIED"], sheet_name])
        cell = index_sheet.cell(row=row_idx, column=5)
        cell.hyperlink = f"#{sheet_name}!A1"
        cell.style = "Hyperlink"

    max_row = index_sheet.max_row
    table = Table(displayName="ScriptsTable", ref=f"A{scripts_header_row + 1}:E{max_row}")
    style = TableStyleInfo(name=INDEX_TABLE_STYLE, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    index_sheet.add_table(table)

    index_sheet.append([])

    index_sheet.append([TABLEAU_HEADING])
    tableau_header_row = index_sheet.max_row
    index_sheet[f"A{tableau_header_row}"].style = header_style
    index_sheet.append(["Data Source", "Field Name", "Calculation", "Data Type"])

    for _, row in tableau_fields.iterrows():
        row_idx = index_sheet.max_row + 1
        index_sheet.append([row["Data Source"], row["Field Name"], row["Calculation"], row["Data Type"]])

    max_row = index_sheet.max_row
    table = Table(displayName="TableauFieldsTable", ref=f"A{tableau_header_row + 1}:D{max_row}")
    style = TableStyleInfo(name=INDEX_TABLE_STYLE, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    index_sheet.add_table(table)

def format_worksheet(ws, table_name, table_schema, header_style, TABLE_STYLE):
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.sheet_view.showGridLines = False

    ws["A1"].style = header_style
    ws["A2"].style = header_style

    max_row = ws.max_row

    table = Table(displayName=table_name, ref=f"A4:C{max_row}")
    style = TableStyleInfo(name=TABLE_STYLE, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

def add_table_to_sheet(wb, table_name, table_info, all_scripts, header_style, TABLE_STYLE):
    table_schema = table_info['schema']
    truncated_name = truncate_table_name(table_name.upper(), wb.sheetnames)

    ws = wb.create_sheet(title=truncated_name)
    ws.append([table_name.upper()])
    ws.append([f"Schema: {table_schema}"])
    ws.append([])
    ws.append(["#", "Column Name", "Data Type"])

    sorted_columns = sorted(table_info['columns'], key=lambda x: x[0])
    for ordinal_position, column_name, data_type in sorted_columns:
        ws.append([str(ordinal_position), column_name, data_type])

    # Calculate the end row for the table based on actual data
    table_end_row = len(sorted_columns) + 4

    # Ensure the table name is unique within the sheet
    base_table_name = f"{truncated_name}_table"
    unique_table_name = base_table_name
    count = 1
    while unique_table_name in [t.name for t in ws._tables]:
        unique_table_name = f"{base_table_name}_{count}"
        count += 1

    # Add the table style
    table = Table(displayName=unique_table_name, ref=f"A4:C{table_end_row}")
    style = TableStyleInfo(name=TABLE_STYLE, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Add the SQL_TRANSFORMATION title
    sql_title_row = 3
    sql_content_start_row = 4
    sql_content_end_row = sql_content_start_row + 499
    ws.cell(row=sql_title_row, column=5, value="SQL_TRANSFORMATION").style = header_style

    # Merge cells for SQL_TRANSFORMATION to span 500 rows and columns E to I
    ws.merge_cells(start_row=sql_content_start_row, start_column=5, end_row=sql_content_end_row, end_column=9)

    # Add the SQL transformation query
    matching_script = all_scripts[all_scripts['DATA_SET_MNEMONIC'] == table_name]
    if not matching_script.empty:
        sql_cell = ws.cell(row=sql_content_start_row, column=5, value=matching_script.iloc[0]['TRANSFORMATION_SQL'])
        sql_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

    # Set the back to index hyperlink
    cell = ws.cell(row=1, column=5, value="Back to Index")
    cell.hyperlink = "#Index!A1"
    cell.style = "Hyperlink"

    # Adjust column widths
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 30

    # Format the worksheet table (only if needed, otherwise skip this step)
    # format_worksheet(ws, truncated_name, table_schema, header_style, TABLE_STYLE)

def add_script_to_sheet(wb, script, index, additional_v_catalog, header_style, TABLE_STYLE):
    sheet_name = f"Script_{index + 1}"
    dataset_mnemonic = script["DATA_SET_MNEMONIC"]
    title = f"{dataset_mnemonic} Script"

    ws = wb.create_sheet(title=sheet_name)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.sheet_view.showGridLines = False

    ws.append([title])
    ws["A1"].style = header_style
    ws.append([])

    ws.append(["WORKFLOW_NAME", script["WORKFLOW_NAME"]])
    ws.append(["DATA_SET_MNEMONIC", script["DATA_SET_MNEMONIC"]])
    ws.append(["DATA_SET_VERSION", str(script["DATA_SET_VERSION"])])
    ws.append(["DATE_MODIFIED", script["DATE_MODIFIED"]])
    ws.append([])
    ws.append(["TRANSFORMATION_SQL"])
    ws["A8"].style = header_style

    start_row = ws.max_row + 1
    end_row = start_row + 500
    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=12)
    sql_cell = ws.cell(row=start_row, column=1, value=script["TRANSFORMATION_SQL"])
    sql_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    sql_cell.font = Font(name='Calibri', size=11)

    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 15

    table_start_row = 8
    table_title_row = table_start_row
    ws.cell(row=table_title_row, column=14, value="Table Columns").style = header_style
    table_start_row += 1
    ws.cell(row=table_start_row, column=14, value="#")
    ws.cell(row=table_start_row, column=15, value="Column Name")
    ws.cell(row=table_start_row, column=16, value="Data Type")

    matching_catalog = additional_v_catalog[(additional_v_catalog['table_name'] == dataset_mnemonic)]
    for _, row in matching_catalog.iterrows():
        table_start_row += 1
        ws.cell(row=table_start_row, column=14, value=row["ordinal_position"])
        ws.cell(row=table_start_row, column=15, value=row["column_name"])
        ws.cell(row=table_start_row, column=16, value=row["data_type"])

    table_end_row = table_start_row
    table = Table(displayName=f"{dataset_mnemonic}_Columns", ref=f"N{table_title_row + 1}:P{table_end_row}")
    style = TableStyleInfo(name=TABLE_STYLE, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    cell = ws.cell(row=1, column=5, value="Back to Index")
    cell.hyperlink = "#Index!A1"
    cell.style = "Hyperlink"

def add_refresh_instructions(wb, header_style):
    # Create a new sheet for refresh instructions and make it the second sheet
    instructions_sheet = wb.create_sheet(title="Refresh Instructions", index=1)
    
    # Remove gridlines
    instructions_sheet.sheet_view.showGridLines = False

    # Write the instructions
    instructions = [
        ("Refresh Instructions", header_style),
        ("This document contains instructions on how to refresh and update the data for this Excel workbook.", None),
        
        ("1. Expected Files:", "Bold"),
        ("   - Two V_CATALOG files:", None),
        ("     1. Specific datasets for the report scope.", None),
        ("     2. Entire Vertica catalog.", None),
        ("   These files are in the V_CATALOG collection on HealtheIntent.", None),
        ("   - Two script files:", None),
        ("     1. Specific dataset in scope of the report.", None),
        ("     2. All scripts.", None),
        ("   These are found in the 'HealtheIntent Transformation Exporter' script in the 'Python Utilities' collection.", None),
        ("", None),
        
        ("2. Running the Datasets:", "Bold"),
        ("   - Run and export these 4 datasets to CSV format.", None),
        ("", None),
        
        ("3. Preparing the Code:", "Bold"),
        ("   - Open Visual Studio Code (VSCode).", None),
        ("   - Pull the code from the repository:", None),
        ("     https://github.com/EddieDavison92/hei-dataset-and-queries-exporter", "Hyperlink"),
        ("", None),
        
        ("4. Saving the Datasets:", "Bold"),
        ("   - Save the four datasets in the 'input' directory.", None),
        ("", None),
        
        ("5. Adding Tableau Workbooks:", "Bold"),
        ("   - Add Tableau workbooks in .TWB format to the 'input/tableau' subfolder.", None),
        ("   - Run 'extract_tableau_calculations.py' to create 'tableau.csv' in the 'output' folder.", None),
        ("", None),
        
        ("6. Modifying Constants and Generating the Excel File:", "Bold"),
        ("   - Adjust constants in the create_catalog functions to load the correct files and update titles/text.", None),
        ("   - Run the script to generate a new Excel file for the project.", None),
        ("", None),
        
        ("If you encounter any issues, please refer to the repository README.", None),
    ]

    for row_num, (text, style) in enumerate(instructions, start=1):
        cell = instructions_sheet.cell(row=row_num, column=1, value=text)
        if style == "Bold":
            cell.font = Font(bold=True)
        elif style == "Hyperlink":
            cell.hyperlink = "https://github.com/EddieDavison92/hei-dataset-and-queries-exporter"
            cell.style = "Hyperlink"
        elif style:
            cell.style = style

def export_to_excel(tables, scripts, tableau_fields, additional_v_catalog, all_scripts, output_path, header_style, INDEX_TABLE_STYLE, TABLE_STYLE, HEADER_INSTRUCTIONS, HEADER_NAVIGATION, HEADER_DATASETS, HEADER_SCRIPTS, TABLEAU_HEADING):
    wb = Workbook()
    wb.remove(wb.active)

    if "header_style" not in wb.named_styles:
        wb.add_named_style(header_style)

    # Add refresh instructions
    add_refresh_instructions(wb, header_style)

    table_names = {}
    for table_name, table_info in tables.items():
        table_names[table_name] = {
            'schema': table_info['schema'],
            'sheet_name': truncate_table_name(table_name.upper(), wb.sheetnames)
        }
        add_table_to_sheet(wb, table_name, table_info, all_scripts, header_style, TABLE_STYLE)

    for i, script in scripts.iterrows():
        add_script_to_sheet(wb, script, i, additional_v_catalog, header_style, TABLE_STYLE)

    create_index_sheet(wb, table_names, scripts, tableau_fields, header_style, INDEX_TABLE_STYLE, HEADER_INSTRUCTIONS, HEADER_NAVIGATION, HEADER_DATASETS, HEADER_SCRIPTS, TABLEAU_HEADING)

    wb.save(output_path)
